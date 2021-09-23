"""
读取excel中的用例数据，并存储到列表中，列表中的每个元素为字典，每一个字典中是标题与一条内容的键值对
"""

from openpyxl import load_workbook
from common.utils import date_format

"""
读取excel中的用例数据，并存储到列表中，列表中的每个元素为字典，每一个字典中是标题与一条内容的键值对
"""
class ReadExcel:
    def __init__(self,filename):
        """
        param filename: excel用例的文件路径
        """
        self.filename = filename
        self.wb = load_workbook(self.filename,data_only=True)
        self.ws = self.wb.active #获取活动工作表
        self.max_column = self.ws.max_column
        self.max_row = self.ws.max_row
        self.case_title = []
        self.case_data = []
        self.case_titledata = []

    def read_excel(self):

        for i in range(1,self.max_column+1):
            '''读取excel从第一列开始，到max_colunmn列结束'''

            self.case_title.append(self.ws.cell(1,i).value)
            # print("标题为：",self.case_title)

        for i in range(2, self.max_row + 1):
            '''读取excel从第二行开始，到max_row行结束，把每行读取到的值以列表的方式存储到casedata列表中'''
            a = []
            for j in range(1, self.max_column + 1):
                a.append(self.ws.cell(i, j).value)
                # print(ws.cell(i,j).value)
            self.case_data.append(a)

        for i in self.case_data:
            '''
            把每行数据通过每列索引取出，以key-value的形式存储到字典，并放到case_titledata大列表中
            再取数据的过程中判断值是否以year=开头，如果是将他处理成相对应的日期存储
            '''
            a = {}
            for j in range(len(self.case_title)):
                if str(i[j]).startswith("year="):
                    '''如果识别出是需要计算的日期，则对应替换为计算出的年月日'''
                    # print("取到的日期值为"+i[j])
                    i[j]=date_format((str(i[j])))
                    a[self.case_title[j]] = i[j]
                elif str(i[j]).startswith("["):
                    '''如果识别出是[开头的，则转换为列表的数据类型'''
                    # print("以[开头的数据为",i[j])
                    # print("以[开头的数据转列表后的类型为",type(eval(i[j])))
                    a[self.case_title[j]] = eval(i[j])

                else:
                    '''试算中除了列表是数据类型，其他都是字符串，统一转换'''
                    a[self.case_title[j]] = str(i[j])
                    # print("非列表与日期为：",str(i[j]))

            self.case_titledata.append(a)
            # print("读取excel后的数据为",self.case_titledata)
        for i in self.case_titledata:
            if i["getResponseValue"]=="$.data.rateTrialRespVO.totalPremium":
                i["assert_value"]=round(float(i["assert_value"]),2)
                if  "." not in str(i["assert_value"]):
                    i["assert_value"] = i["assert_value"] +".0"


        # print("读取excel后的数据为",self.case_titledata)
        return self.case_titledata

if __name__ == '__main__':
    from common import contants
    read_excel = ReadExcel(contants.case_zhengqingchun).read_excel()